"""
tracker/exporter.py  —  Excel and CSV export

Sheets:
  1. Summary        — current holdings with P&L
  2. Transactions   — full transaction history with commissions
  3. Dividends      — all dividend payments with per-year subtotals
  4. FIFO Gains     — realised gains/losses per sell (FIFO cost basis)
  5. Tax Summary    — German Abgeltungsteuer estimate per year
"""

import csv
from datetime import datetime
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from tracker.models import Dividend, Holding

# ── Colour palette ─────────────────────────────────────────────────────────────
HEADER_BG  = "1A237E"
HEADER_FG  = "FFFFFF"
SUBHEAD_BG = "283593"
POS_FG     = "1B5E20"
NEG_FG     = "B71C1C"
ALT_ROW    = "E8EAF6"
DIV_BG     = "E8F5E9"   # light green for dividend rows
TAX_BG     = "FFF8E1"   # light amber for tax rows
GAIN_BG    = "E8F5E9"
LOSS_BG    = "FFEBEE"

def _side():        return Side(style="thin", color="BDBDBD")
def _border():      return Border(left=_side(), right=_side(), top=_side(), bottom=_side())
def _hfont(bold=True, size=10, color=HEADER_FG):
    return Font(name="Arial", size=size, bold=bold, color=color)
def _hfill(bg=HEADER_BG): return PatternFill("solid", fgColor=bg)
def _dfont(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def _cell(ws, row, col, value=None, font=None, fill=None, fmt=None,
          align="left", bold=False, color="000000"):
    c = ws.cell(row, col)
    if value is not None: c.value = value
    c.font      = font or _dfont(bold=bold, color=color)
    if fill:    c.fill = fill
    if fmt:     c.number_format = fmt
    c.border    = _border()
    c.alignment = Alignment(horizontal=align)
    return c

def _title_row(ws, row, text, cols, size=14, bg=HEADER_BG):
    ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")
    c = ws.cell(row, 1, text)
    c.font      = Font(name="Arial", size=size, bold=True, color=HEADER_FG)
    c.fill      = _hfill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24

def _header_row(ws, row, headers, bg=HEADER_BG):
    for col, h in enumerate(headers, 1):
        _cell(ws, row, col, h, font=_hfont(), fill=_hfill(bg), align="center")
    ws.row_dimensions[row].height = 18

def _subtotal_row(ws, row, cols, labels: dict, bg=SUBHEAD_BG):
    """labels = {col_index: value_or_formula}"""
    for col in range(1, cols + 1):
        c = ws.cell(row, col)
        c.fill   = _hfill(bg)
        c.border = _border()
        if col in labels:
            c.value     = labels[col]
            c.font      = _hfont()
            c.alignment = Alignment(horizontal="right" if col > 2 else "left")


# ── Public API ─────────────────────────────────────────────────────────────────

def export_to_excel(holdings: List[Holding],
                    prices: Dict[str, Optional[float]],
                    dividends: Optional[List[Dividend]] = None,
                    filename: Optional[str] = None) -> str:
    if filename is None:
        filename = f"portfolio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    if dividends is None:
        dividends = []

    wb = openpyxl.Workbook()
    _summary_sheet(wb, holdings, prices)
    _transactions_sheet(wb, holdings)
    _dividends_sheet(wb, dividends, holdings)
    _fifo_gains_sheet(wb, holdings)
    _tax_summary_sheet(wb, holdings, dividends)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(filename)
    return filename


def export_to_csv(holdings: List[Holding],
                  prices: Dict[str, Optional[float]],
                  filename: Optional[str] = None) -> str:
    if filename is None:
        filename = f"portfolio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    fields = ["ticker","name","type","quantity","avg_cost",
              "current_price","market_value","unrealised_pnl","pnl_percent"]
    with open(filename, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for h in holdings:
            p = prices.get(h.ticker)
            w.writerow({
                "ticker":        h.ticker,
                "name":          h.name,
                "type":          h.asset_type,
                "quantity":      round(h.quantity, 6),
                "avg_cost":      round(h.average_cost, 4),
                "current_price": round(p, 4) if p else "N/A",
                "market_value":  round(h.current_value(p), 2) if p else "N/A",
                "unrealised_pnl":round(h.unrealised_pnl(p), 2) if p else "N/A",
                "pnl_percent":   round(h.pnl_percent(p), 2) if p else "N/A",
            })
    return filename


# ── Sheet 1: Summary ──────────────────────────────────────────────────────────

def _summary_sheet(wb, holdings, prices):
    ws = wb.create_sheet("Summary")
    _title_row(ws, 1, f"Portfolio Summary  —  {datetime.now().strftime('%d %b %Y  %H:%M')}", 9)

    headers = ["Ticker","Name","Type","Quantity",
               "Avg Cost (€)","Current Price (€)","Market Value (€)",
               "Unrealised P&L (€)","P&L %"]
    _header_row(ws, 3, headers)

    data_start = 4
    for i, h in enumerate(holdings):
        row  = data_start + i
        p    = prices.get(h.ticker)
        fill = _hfill(ALT_ROW) if i % 2 == 0 else None

        _cell(ws, row, 1, h.ticker,          fill=fill)
        _cell(ws, row, 2, h.name,            fill=fill)
        _cell(ws, row, 3, h.asset_type.upper(), fill=fill)
        _cell(ws, row, 4, h.quantity,        fill=fill, fmt="#,##0.0000", align="right")
        _cell(ws, row, 5, h.average_cost,    fill=fill, fmt='€#,##0.00;(€#,##0.00);"-"', align="right")
        _cell(ws, row, 6, p if p else "N/A", fill=fill, fmt='€#,##0.00;(€#,##0.00);"-"', align="right")

        if p:
            mv_col   = get_column_letter(7)
            cost_col = get_column_letter(5)
            qty_col  = get_column_letter(4)
            # Market value = qty × price (formula)
            ws.cell(row, 7).value          = f"=D{row}*F{row}"
            ws.cell(row, 7).number_format  = '€#,##0.00;(€#,##0.00);"-"'
            ws.cell(row, 7).font           = _dfont()
            ws.cell(row, 7).fill           = fill or PatternFill()
            ws.cell(row, 7).border         = _border()
            ws.cell(row, 7).alignment      = Alignment(horizontal="right")
            # P&L = market value - (avg_cost × qty)
            ws.cell(row, 8).value          = f"=G{row}-(E{row}*D{row})"
            pnl_val = h.unrealised_pnl(p)
            pnl_color = POS_FG if pnl_val >= 0 else NEG_FG
            ws.cell(row, 8).number_format  = '€#,##0.00;[Red](€#,##0.00);"-"'
            ws.cell(row, 8).font           = _dfont(color=pnl_color)
            ws.cell(row, 8).fill           = fill or PatternFill()
            ws.cell(row, 8).border         = _border()
            ws.cell(row, 8).alignment      = Alignment(horizontal="right")
            # P&L % = P&L / cost basis
            ws.cell(row, 9).value          = f"=IF(E{row}*D{row}<>0,H{row}/(E{row}*D{row}),0)"
            ws.cell(row, 9).number_format  = '0.00%;[Red]-0.00%;"-"'
            ws.cell(row, 9).font           = _dfont(color=pnl_color)
            ws.cell(row, 9).fill           = fill or PatternFill()
            ws.cell(row, 9).border         = _border()
            ws.cell(row, 9).alignment      = Alignment(horizontal="right")
        else:
            for col in (7, 8, 9):
                _cell(ws, row, col, "N/A", fill=fill, align="right")

    # Totals row
    tr = data_start + len(holdings)
    n  = len(holdings)
    _subtotal_row(ws, tr, 9, {
        1: "TOTAL",
        7: f"=SUM(G{data_start}:G{tr-1})",
        8: f"=SUM(H{data_start}:H{tr-1})",
        9: f"=IF(SUMPRODUCT(E{data_start}:E{tr-1},D{data_start}:D{tr-1})<>0,"
           f"H{tr}/SUMPRODUCT(E{data_start}:E{tr-1},D{data_start}:D{tr-1}),0)",
    })
    ws.cell(tr, 7).number_format = '€#,##0.00;(€#,##0.00);"-"'
    ws.cell(tr, 8).number_format = '€#,##0.00;[Red](€#,##0.00);"-"'
    ws.cell(tr, 9).number_format = '0.00%;[Red]-0.00%;"-"'

    for i, w in enumerate([10,24,8,12,16,18,18,20,10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


# ── Sheet 2: Transactions ─────────────────────────────────────────────────────

def _transactions_sheet(wb, holdings):
    ws = wb.create_sheet("Transactions")
    headers = ["Ticker","Name","Date","Action",
               "Quantity","Price (€)","Commission (€)","Total Cost (€)"]
    _title_row(ws, 1, "Transaction History", len(headers))
    _header_row(ws, 3, headers)

    row = 4
    all_txns = sorted(
        [(t.date, h.ticker, h.name, t) for h in holdings for t in h.transactions],
        key=lambda x: x[0]
    )
    for i, (_, ticker, name, t) in enumerate(all_txns):
        fill = _hfill("E8F5E9") if t.action == "buy" else _hfill("FFEBEE")
        _cell(ws, row, 1, ticker,              fill=fill)
        _cell(ws, row, 2, name,                fill=fill)
        _cell(ws, row, 3, t.date,              fill=fill)
        _cell(ws, row, 4, t.action.upper(),    fill=fill, align="center")
        _cell(ws, row, 5, t.quantity,          fill=fill, fmt="#,##0.0000",            align="right")
        _cell(ws, row, 6, t.price,             fill=fill, fmt='€#,##0.00;(€#,##0.00)', align="right")
        _cell(ws, row, 7, t.commission,        fill=fill, fmt='€#,##0.00;(€#,##0.00)', align="right")
        # Total cost formula: qty × price + commission
        ws.cell(row, 8).value         = f"=E{row}*F{row}+G{row}"
        ws.cell(row, 8).number_format = '€#,##0.00;(€#,##0.00)'
        ws.cell(row, 8).font          = _dfont()
        ws.cell(row, 8).fill          = fill
        ws.cell(row, 8).border        = _border()
        ws.cell(row, 8).alignment     = Alignment(horizontal="right")
        row += 1

    # Totals
    if row > 4:
        _subtotal_row(ws, row, 8, {
            1: "TOTAL",
            7: f"=SUM(G4:G{row-1})",
            8: f"=SUM(H4:H{row-1})",
        })
        ws.cell(row, 7).number_format = '€#,##0.00;(€#,##0.00)'
        ws.cell(row, 8).number_format = '€#,##0.00;(€#,##0.00)'

    for i, w in enumerate([10,24,12,8,12,14,16,16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


# ── Sheet 3: Dividends ────────────────────────────────────────────────────────

def _dividends_sheet(wb, dividends: List[Dividend], holdings: List[Holding]):
    ws = wb.create_sheet("Dividends")
    headers = ["Date","Ticker","Name","Gross Amount (€)","Withholding Tax (€)","Net Amount (€)"]
    _title_row(ws, 1, "Dividend Income", len(headers))
    _header_row(ws, 3, headers)

    if not dividends:
        ws.merge_cells("A4:F4")
        ws.cell(4, 1).value     = "No dividends recorded."
        ws.cell(4, 1).font      = _dfont(color="888888")
        ws.cell(4, 1).alignment = Alignment(horizontal="center")
        return

    name_map = {h.ticker: h.name for h in holdings}
    sorted_divs = sorted(dividends, key=lambda d: d.date)

    row = 4
    by_year: Dict[str, list] = {}
    for d in sorted_divs:
        y = d.date[:4]
        by_year.setdefault(y, []).append(row)

        fill = _hfill(DIV_BG) if row % 2 == 0 else None
        _cell(ws, row, 1, d.date,                        fill=fill)
        _cell(ws, row, 2, d.ticker,                      fill=fill)
        _cell(ws, row, 3, name_map.get(d.ticker, ""),    fill=fill)
        _cell(ws, row, 4, d.amount,                      fill=fill, fmt='€#,##0.00;(€#,##0.00)', align="right")
        _cell(ws, row, 5, d.withholding_tax,             fill=fill, fmt='€#,##0.00;(€#,##0.00)', align="right")
        ws.cell(row, 6).value         = f"=D{row}-E{row}"
        ws.cell(row, 6).number_format = '€#,##0.00;(€#,##0.00)'
        ws.cell(row, 6).font          = _dfont()
        ws.cell(row, 6).fill          = fill or PatternFill()
        ws.cell(row, 6).border        = _border()
        ws.cell(row, 6).alignment     = Alignment(horizontal="right")
        row += 1

    # Per-year subtotals
    for year, rows in sorted(by_year.items()):
        first, last = rows[0], rows[-1]
        _subtotal_row(ws, row, 6, {
            1: f"{year} Total",
            4: f"=SUM(D{first}:D{last})",
            5: f"=SUM(E{first}:E{last})",
            6: f"=SUM(F{first}:F{last})",
        })
        for col in (4, 5, 6):
            ws.cell(row, col).number_format = '€#,##0.00;(€#,##0.00)'
        row += 1

    # Grand total
    _subtotal_row(ws, row, 6, {
        1: "GRAND TOTAL",
        4: f"=SUM(D4:D{row-1})",
        5: f"=SUM(E4:E{row-1})",
        6: f"=SUM(F4:F{row-1})",
    })
    for col in (4, 5, 6):
        ws.cell(row, col).number_format = '€#,##0.00;(€#,##0.00)'

    for i, w in enumerate([12,10,24,18,20,16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


# ── Sheet 4: FIFO Gains ───────────────────────────────────────────────────────

def _fifo_gains_sheet(wb, holdings: List[Holding]):
    from tracker.tax import compute_realised_gains
    ws = wb.create_sheet("FIFO Gains")
    headers = ["Sell Date","Ticker","Qty Sold","Proceeds (€)",
               "Sell Commission (€)","FIFO Cost Basis (€)","Gain / Loss (€)"]
    _title_row(ws, 1, "Realised Gains & Losses (FIFO)", len(headers))
    _header_row(ws, 3, headers)

    gains = compute_realised_gains({h.ticker: h for h in holdings})

    if not gains:
        ws.merge_cells("A4:G4")
        ws.cell(4, 1).value     = "No sell transactions recorded."
        ws.cell(4, 1).font      = _dfont(color="888888")
        ws.cell(4, 1).alignment = Alignment(horizontal="center")
        return

    row = 4
    for g in gains:
        is_gain = g.gain >= 0
        fill    = _hfill(GAIN_BG) if is_gain else _hfill(LOSS_BG)
        pnl_col = POS_FG if is_gain else NEG_FG

        _cell(ws, row, 1, g.sell_date,   fill=fill)
        _cell(ws, row, 2, g.ticker,      fill=fill)
        _cell(ws, row, 3, g.quantity,    fill=fill, fmt="#,##0.0000",              align="right")
        _cell(ws, row, 4, g.proceeds,    fill=fill, fmt='€#,##0.00;(€#,##0.00)',   align="right")
        _cell(ws, row, 5, g.commission,  fill=fill, fmt='€#,##0.00;(€#,##0.00)',   align="right")
        _cell(ws, row, 6, g.cost_basis,  fill=fill, fmt='€#,##0.00;(€#,##0.00)',   align="right")
        # Gain = proceeds - commission - cost basis (formula)
        ws.cell(row, 7).value         = f"=D{row}-E{row}-F{row}"
        ws.cell(row, 7).number_format = '€#,##0.00;[Red](€#,##0.00);"-"'
        ws.cell(row, 7).font          = _dfont(color=pnl_col)
        ws.cell(row, 7).fill          = fill
        ws.cell(row, 7).border        = _border()
        ws.cell(row, 7).alignment     = Alignment(horizontal="right")
        row += 1

    # Totals
    _subtotal_row(ws, row, 7, {
        1: "TOTAL",
        4: f"=SUM(D4:D{row-1})",
        5: f"=SUM(E4:E{row-1})",
        6: f"=SUM(F4:F{row-1})",
        7: f"=SUM(G4:G{row-1})",
    })
    for col in (4, 5, 6, 7):
        ws.cell(row, col).number_format = '€#,##0.00;[Red](€#,##0.00);"-"'

    for i, w in enumerate([12,10,12,16,20,18,16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


# ── Sheet 5: Tax Summary ──────────────────────────────────────────────────────

def _tax_summary_sheet(wb, holdings: List[Holding], dividends: List[Dividend]):
    from tracker.tax import year_summary, all_active_years, SPARERPAUSCHBETRAG
    ws = wb.create_sheet("Tax Summary (DE)")
    headers = [
        "Year",
        "Realised Gains (€)","Realised Losses (€)","Dividend Income (€)",
        "Net Taxable (€)","Sparerpauschbetrag (€)","Taxable Base (€)",
        "Abgeltungsteuer 25% (€)","Solidaritätszuschlag (€)",
        "Withholding Credit (€)","Est. Tax Owed (€)",
        "Allowance Remaining (€)"
    ]
    _title_row(ws, 1, "German Tax Summary — Abgeltungsteuer Estimate", len(headers))

    # Disclaimer
    ws.merge_cells(f"A2:{get_column_letter(len(headers))}2")
    disc = ws.cell(2, 1, "⚠  Estimate only. Kirchensteuer not included. Consult a Steuerberater for official filing.")
    disc.font      = Font(name="Arial", size=9, italic=True, color="B71C1C")
    disc.alignment = Alignment(horizontal="center")

    _header_row(ws, 4, headers)

    holdings_dict = {h.ticker: h for h in holdings}
    years = all_active_years(holdings_dict, dividends)
    if not years:
        ws.cell(5, 1).value = "No transactions recorded."
        return

    row = 5
    for year in years:
        s    = year_summary(year, holdings_dict, dividends)
        fill = _hfill(TAX_BG) if row % 2 == 0 else None

        _cell(ws, row, 1,  str(year),              fill=fill, bold=True)
        _cell(ws, row, 2,  s.realised_gains,        fill=fill, fmt='€#,##0.00;(€#,##0.00)', align="right")
        _cell(ws, row, 3,  s.realised_losses,       fill=fill, fmt='€#,##0.00;(€#,##0.00)', align="right")
        _cell(ws, row, 4,  s.dividend_income,       fill=fill, fmt='€#,##0.00;(€#,##0.00)', align="right")
        # Net taxable = gains - losses + dividends (formula)
        ws.cell(row, 5).value         = f"=B{row}-C{row}+D{row}"
        ws.cell(row, 5).number_format = '€#,##0.00;(€#,##0.00)'
        ws.cell(row, 5).font          = _dfont()
        ws.cell(row, 5).fill          = fill or PatternFill()
        ws.cell(row, 5).border        = _border()
        ws.cell(row, 5).alignment     = Alignment(horizontal="right")

        _cell(ws, row, 6,  min(SPARERPAUSCHBETRAG, s.net_taxable),
              fill=fill, fmt='€#,##0.00;(€#,##0.00)', align="right")
        # Taxable base = MAX(0, net_taxable - allowance)
        ws.cell(row, 7).value         = f"=MAX(0,E{row}-F{row})"
        ws.cell(row, 7).number_format = '€#,##0.00;(€#,##0.00)'
        ws.cell(row, 7).font          = _dfont()
        ws.cell(row, 7).fill          = fill or PatternFill()
        ws.cell(row, 7).border        = _border()
        ws.cell(row, 7).alignment     = Alignment(horizontal="right")
        # Abgeltungsteuer = 25%
        ws.cell(row, 8).value         = f"=G{row}*0.25"
        ws.cell(row, 8).number_format = '€#,##0.00;(€#,##0.00)'
        ws.cell(row, 8).font          = _dfont()
        ws.cell(row, 8).fill          = fill or PatternFill()
        ws.cell(row, 8).border        = _border()
        ws.cell(row, 8).alignment     = Alignment(horizontal="right")
        # Soli = 5.5% of Abgeltungsteuer
        ws.cell(row, 9).value         = f"=H{row}*0.055"
        ws.cell(row, 9).number_format = '€#,##0.00;(€#,##0.00)'
        ws.cell(row, 9).font          = _dfont()
        ws.cell(row, 9).fill          = fill or PatternFill()
        ws.cell(row, 9).border        = _border()
        ws.cell(row, 9).alignment     = Alignment(horizontal="right")

        _cell(ws, row, 10, s.withholding_credit,
              fill=fill, fmt='€#,##0.00;(€#,##0.00)', align="right")
        # Est. tax owed = MAX(0, Abgelt + Soli - WHT credit)
        ws.cell(row, 11).value         = f"=MAX(0,H{row}+I{row}-J{row})"
        ws.cell(row, 11).number_format = '€#,##0.00;(€#,##0.00)'
        ws.cell(row, 11).font          = _dfont(bold=True, color=NEG_FG)
        ws.cell(row, 11).fill          = fill or PatternFill()
        ws.cell(row, 11).border        = _border()
        ws.cell(row, 11).alignment     = Alignment(horizontal="right")

        _cell(ws, row, 12, s.allowance_remaining,
              fill=fill, fmt='€#,##0.00;(€#,##0.00)', align="right")
        row += 1

    for i, w in enumerate([8,18,18,18,16,22,16,22,22,20,16,22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
